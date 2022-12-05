import sys

import pandas
import os
import json
from openpyxl import load_workbook
from urllib import parse
import pathlib
import html

logo = os.path.abspath('TTLFinalLogo.jpg')

FAILED_IMAGES_JSON = os.path.join("chefdata", "failed_links", "failed_image_links.json")

DICT_IMAGES = {'Mathematics': 'TTL_math_practice_content',
               'Science': 'TTL_science_practice_content'}


def integer_to_roman(num):
    # Storing roman values of digits from 0-9
    # when placed at different places
    m = ["", "M", "MM", "MMM"]
    c = ["", "C", "CC", "CCC", "CD", "D",
         "DC", "DCC", "DCCC", "CM "]
    x = ["", "X", "XX", "XXX", "XL", "L",
         "LX", "LXX", "LXXX", "XC"]
    i = ["", "I", "II", "III", "IV", "V",
         "VI", "VII", "VIII", "IX"]

    # Converting to roman
    thousands = m[num // 1000]
    hundreds = c[(num % 1000) // 100]
    tens = x[(num % 100) // 10]
    ones = i[num % 10]

    ans = (thousands + hundreds +
           tens + ones)

    return ans


def read_videos_xls(xls):
    """
    Build dict for channel structure
    Dict structure:
        Language > Grade > Subject > Chapter > Topic Name > Content Type > Content
    """
    wb = load_workbook(xls)
    sheetnames = wb.sheetnames
    dict_sheet_names = {}
    for sheet_name in sheetnames:
        if 'english' in sheet_name.lower():
            if 'english' not in dict_sheet_names:
                dict_sheet_names['english'] = [sheet_name]
            else:
                dict_sheet_names['english'].append(sheet_name)

    # building dict row by row
    data_dict = {}
    for language in dict_sheet_names:
        for sheet in dict_sheet_names[language]:
            data_from_xls = pandas.read_excel(xls, sheet_name=sheet, keep_default_na=False, na_values='')
            for index, row in data_from_xls.iterrows():
                # Skip any content that does not have a link and all assessments
                # as we will be getting assessment data from a separate xls
                # adding language keys to dict
                # converting strings to lower due to inconsistent naming convention in xls
                # language = row["Language"].lower().strip()
                if language not in data_dict:
                    data_dict[language] = {}

                # addingggrade keys to dict
                grade = row["Grade"]
                if grade not in data_dict[language]:
                    data_dict[language][grade] = {}

                subject = sheet.split(' ')[0]
                if subject not in data_dict[language][grade]:
                    data_dict[language][grade][subject] = {}

                chapter_num = row['Chapter Number']
                chapter_name = row["Chapter Name"].lower().strip()
                chapter_name = chapter_name.replace('?', '')
                chapter = "{} - {}".format(chapter_num, chapter_name)

                if chapter not in data_dict[language][grade][subject]:
                    data_dict[language][grade][subject][chapter] = {}

                topic_name = row["Topic Name"].lower().strip()
                if topic_name not in data_dict[language][grade][subject][chapter]:
                    data_dict[language][grade][subject][chapter][topic_name] = {}
                # adding chapter assessment
                # data_dict[language][grade][subject][chapter][topic_name]["chapter assessment"] = {}

                # Assessment = Exercise, Video = Video
                content_type = 'Video'
                if content_type not in data_dict[language][grade][subject][chapter][topic_name]:
                    data_dict[language][grade][subject][chapter][topic_name][content_type] = {}
                if row.get("Branded video link"):
                    link = row["Branded video link"].lower().strip()
                elif row.get("Branded video"):
                    link = row["Branded video"].lower().strip()
                else:
                    link = row.get("Link to Content").lower().strip()
                    if 'dropbox' in link.lower():
                        print()

                if link not in data_dict[language][grade][subject][chapter][topic_name][content_type]:
                    data_dict[language][grade][subject][chapter][topic_name][content_type][link] = {
                        "title": row["Video topic as per Youtube"],
                        "copyright": 'TicTacLearn',
                        "license": 'TicTacLearn',
                        "icon": logo
                    }
    return data_dict


def read_assessment_xls(dict_xls, data):
    for key in dict_xls:
        data_from_xls = pandas.read_excel(dict_xls.get(key), keep_default_na=False, na_values='', engine='openpyxl')
        # to map to correct option given which is the right answer
        for index, row in data_from_xls.iterrows():
            question_parts = [x.strip() for x in row["Question Set Name"].split("|")]
            topic = None
            chapter_title = None
            if len(question_parts) == 4:
                chapter_title = question_parts[0]
            elif len(question_parts) == 5:
                # if normal assessment, chapter is located at str_parts[1] and topic is located at str_parts[0]
                topic = question_parts[0].lower().strip()
                chapter_title = question_parts[1].strip()
            elif len(question_parts) == 7 or len(question_parts) == 6:
                # if normal assessment, chapter is located at str_parts[1] and topic is located at str_parts[0]
                topic = question_parts[0].lower().strip()
                chapter_title = question_parts[1].strip()
            else:
                print("Error: Question Set Name not in correct format")
                # self.add_to_failed(row["Medium"], row["Question Set Name"], row["QuestionText"])
                continue
            language = row["Medium"]
            grade = row["Class"]
            grade = "Grade_{}".format(integer_to_roman(grade))

            if row["Subject"] == "Math" or row["Subject"] == "Maths":
                # provided xls from TTL has Subject listed as mathematics on videos.xls and math/maths in assessments.xls
                subject = 'Mathematics'
            else:
                subject = row["Subject"].capitalize()

            chapter = "Chapter_{}_{}".format(row["ChapterNo"], chapter_title.replace(' ', '_').upper())
            image_path = os.path.abspath(os.path.join(DICT_IMAGES.get(subject), language))
            content_type = "assessment"

            # some questions only have an image with no text
            question_text = row["QuestionText"] if not pandas.isnull(row["QuestionText"]) else None

            question_id = row["QuestionId"]

            question_image = None  # None if no associate question image
            # will remain None if there is no option (ex: yes or no question will only have 2 options)
            option_one = None
            option_two = None
            option_three = None
            option_four = None
            # holds all available answers
            answers = []
            # query which column the options are located in and if there is a question Image
            if not pandas.isnull(row["QuestionImage"]):
                question_image = "![]({})".format(get_image_path(image_path, row["QuestionImage"]))

            if (not pandas.isnull(row["Option1"])) or (not pandas.isnull(row["Option1Image"])):
                option_one = row["Option1"] if not pandas.isnull(row["Option1"]) else "![]({})".format(get_image_path(
                    image_path, row["Option1Image"]))
                answers.append(option_one)

            if (not pandas.isnull(row["Option2"])) or (not pandas.isnull(row["Option2Image"])):
                option_two = row["Option2"] if not pandas.isnull(row["Option2"]) else "![]({})".format(get_image_path(
                    image_path, row["Option2Image"]))
                answers.append(option_two)

            if (not pandas.isnull(row["Option3"])) or (not pandas.isnull(row["Option3Image"])):
                option_three = row["Option3"] if not pandas.isnull(row["Option3"]) else "![]({})".format(get_image_path(
                    image_path, row["Option3Image"]))
                answers.append(option_three)

            if (not pandas.isnull(row["Option4"])) or (not pandas.isnull(row["Option4Image"])):
                option_four = row["Option4"] if not pandas.isnull(row["Option4"]) else "![]({})".format(get_image_path(
                    image_path, row["Option4Image"]))
                answers.append(option_four)

            answer_key = {
                1: option_one,
                2: option_two,
                3: option_three,
                4: option_four,
            }

            answer = answer_key[row["AnswerNo"]]

            if len(answers) > 4:
                print(answers)

            question_metadata = {
                "question": question_text,
                "question_image": question_image,
                "correct_answer": answer,
                "all_answers": answers
            }

            # check if chapter exists as some chapters only appear on assessment excel
            if data and data.get(language) \
                    and data[language].get(grade) \
                    and data[language][grade].get(subject) \
                    and chapter not in data[language][grade][subject]:
                data[language][grade][subject][chapter] = {}

            # have to check if topic exists
            if topic:
                if topic not in data[language][grade][subject][chapter]:
                    data[language][grade][subject][chapter][topic] = {}
                    # add in assessment array. array will hold each question's metadata
                if content_type not in data[language][grade][subject][chapter][topic]:
                    data[language][grade][subject][chapter][topic][content_type] = {}
                data[language][grade][subject][chapter][topic][content_type][question_id] = question_metadata
                # print(data[language][grade][subject][chapter][topic][content_type])
            else:
                chapter_assessment = "Chapter Assessment"
                if data.get(language) \
                        and data[language].get(grade) \
                        and data[language][grade].get(subject) \
                        and data[language][grade][subject].get(chapter) \
                        and chapter_assessment not in data[language][grade][subject][chapter]:
                    data[language][grade][subject][chapter][chapter_assessment] = {}
                if data[language][grade].get(subject) and data[language][grade][subject].get(chapter) and \
                        data[language][grade][subject][chapter].get(chapter_assessment):
                    data[language][grade][subject][chapter][chapter_assessment][question_id] = question_metadata
            # location where to add question objects

    return data


def get_image_path(image_path, image_string):
    # path_arr = image_string.replace(" ", "").split('/')
    path = os.path.join(image_path, image_string)
    if not os.path.isfile(path):
        add_to_failed(image_path)
    return path


def add_to_failed(path_arr):
    grade = path_arr[1]
    chapter = path_arr[2]
    image_name = path_arr[3]
    with open(FAILED_IMAGES_JSON, encoding='utf-8') as f:
        try:
            data = json.loads(f.read())
        except:
            # no data in json file
            print('no data in json file')
            with open(FAILED_IMAGES_JSON, 'w', encoding='utf-8') as json_file:
                dict_failed = {}
                dict_failed[grade] = {}
                dict_failed[grade][chapter] = []
                dict_failed[grade][chapter].append(image_name)
                json.dump(dict_failed, json_file, indent=4, ensure_ascii=False)
                return

        with open(FAILED_IMAGES_JSON, 'w', encoding='utf-8') as json_file:
            if grade not in data:
                data[grade] = {}
            if chapter not in data[grade]:
                data[grade][chapter] = []
            data[grade][chapter].append(image_name)
            json.dump(data, json_file, indent=4, ensure_ascii=False)
            return


def make_the_correct_path(language, subject, grade_string, chapter, vt, video_name=None):
    """
    folder_path: Content from this folder path need to be check
    name: Looking if this name is inside this folder path
    """
    if video_name:
        lst_folders = [language, subject, grade_string, chapter, vt, video_name]
    else:
        lst_folders = [language, subject, grade_string, chapter, vt]
    folder_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads')

    for folder in lst_folders:
        tmp_folder_path = os.path.join(os.path.join(folder_path, folder))
        pathlib.Path(tmp_folder_path).exists()
        if os.path.exists(tmp_folder_path):
            folder_path = tmp_folder_path
            continue
        else:
            lst_content = os.listdir(folder_path)
            for content in lst_content:
                if folder in content:
                    folder_path = os.path.join(os.path.join(folder_path, content))
                    break

    if os.path.isfile(folder_path):
        return folder_path
    return None


def get_all_local_files(xls, language):
    dict_all_files_with_local_path = {}
    dict_sheet_names = {}
    # sheet_name = 'Math (1-10) English' # TESTING REASON

    excel_file = pandas.ExcelFile(xls, engine='openpyxl')
    for sheet_name in excel_file.sheet_names:
        splitted = sheet_name.split(' ')
        if splitted[-1] == language:
            dict_sheet_names[sheet_name] = {'subject': splitted[0], 'language': splitted[-1]}
    # to map all the sheets in the given excel
    for sheet_name in dict_sheet_names:
        data_from_xls = pandas.read_excel(xls, keep_default_na=False, na_values='', sheet_name=[sheet_name],
                                          engine='openpyxl')
        # to map to correct option given which is the right answer
        for index, row in data_from_xls.get(sheet_name).iterrows():
            language = sheet_name.split(' ')[-1]
            subject = sheet_name.split(' ')[0]
            if subject == 'Math':
                subject = 'Mathematics'
            grade = row['Grade']
            grade_string = 'Grade_{}'.format(grade)
            chapter_number = row.get('Chapter No')
            if chapter_number is None:
                chapter_number = row.get('Chapter Number')
            chapter_name = str(row['Chapter Name']).replace('?', '_')
            chapter = None
            try:
                chapter = 'Chapter_{}_{}'.format(int(chapter_number), chapter_name.strip().replace(' ', '_').upper())
            except Exception as ex:
                print(chapter_name)
                print(ex)
            vt_number = row.get('Video Topic Number ')
            try:
                if vt_number is None:
                    vt_number = row.get('Video Topic Number')
                elif vt_number is None:
                    vt_number = (row.get('No of videos in the VT'))
                if vt_number:
                    vt_number = str(int(vt_number))
            except Exception as ex:
                print(vt_number)
                print(ex)
            vt_name = str(row['Topic Name']).lower().strip()
            vt_name = vt_name.replace('?', '_')
            vt = 'VT_{}_{}'.format(vt_number, vt_name.strip().replace(' ', '_').upper())
            vt_path = make_the_correct_path(language, subject, grade_string, chapter, vt)
            if vt_path:
                vt = vt_path.split(os.path.sep)[-1]

            video_name = str(row.get('Branded video link') or row.get('Branded video'))
            video_name = parse.unquote(video_name.split('/')[-1].split('?')[0])
            content_type = 'video'
            file_path = None
            if chapter:
                file_path = make_the_correct_path(language, subject, grade_string, chapter, vt, video_name)

            if language not in dict_all_files_with_local_path:
                dict_all_files_with_local_path[language] = {}
            if grade_string not in dict_all_files_with_local_path[language]:
                dict_all_files_with_local_path[language][grade_string] = {}
            if subject not in dict_all_files_with_local_path[language][grade_string]:
                dict_all_files_with_local_path[language][grade_string][subject] = {}
            if chapter not in dict_all_files_with_local_path[language][grade_string][subject]:
                dict_all_files_with_local_path[language][grade_string][subject][chapter] = {}
            if vt_name not in dict_all_files_with_local_path[language][grade_string][subject][chapter]:
                dict_all_files_with_local_path[language][grade_string][subject][chapter][vt_name] = {}
            if content_type not in dict_all_files_with_local_path[language][grade_string][subject][chapter][vt_name]:
                dict_all_files_with_local_path[language][grade_string][subject][chapter][vt_name][content_type] = {}
            if video_name not in dict_all_files_with_local_path[language][grade_string][subject][chapter][vt_name][
                content_type] and file_path:
                dict_all_files_with_local_path[language][grade_string][subject][chapter][vt_name][content_type][
                    video_name] = str(file_path)
    return dict_all_files_with_local_path
